import openpyxl
import csv
import os

# Define target wages mapping by Resource ID
NEW_WAGES = {
    1: 120.00,  # G.C. General Management
    2: 95.00,   # G.C. Project Management
    3: 75.00,   # G.C. Procurement
    4: 80.00,   # G.C. Scheduler
    5: 70.00,   # G.C. Accounting
    6: 85.00,   # G.C. Superintendent
    7: 65.00,   # G.C. Survey Crew
    8: 55.00,   # G.C. Rough Carpenter Crew
    9: 30.00,   # G.C. Labor Crew
    10: 60.00,  # G.C. Concrete Crew
    11: 60.00,  # G.C. Finish Carpenter Crew
    12: 70.00,  # Site Grading Contractor
    13: 65.00,  # Plumbing Contractor
    14: 80.00,  # Plumbing Contractor Management
    15: 65.00,  # Electric Contractor
    16: 80.00,  # Electric Contractor Management
    17: 65.00,  # HVAC Contractor
    18: 80.00,  # HVAC Contractor Management
    19: 80.00,  # Elevator Contractor
    20: 90.00,  # Elevator Contractor Management
    21: 75.00,  # Steel Erection Contractor
    22: 85.00,  # Steel Erection Contractor Management
    23: 50.00,  # Drywall Contractor
    24: 55.00,  # Masonry Contractor
    25: 55.00,  # Tile Contractor
    26: 55.00,  # Roofing Contractor
    27: 75.00,  # Roofing Contractor Management
    28: 50.00,  # Window Contractor
    29: 45.00,  # Carpet Contractor
    30: 45.00,  # Landscape Contractor
    31: 60.00,  # Paving Contractor
    32: 45.00,  # Painting Contractor
}

def update_xlsx(filepath):
    print(f"Updating Excel file: {filepath}")
    wb = openpyxl.load_workbook(filepath)
    if 'Resource_Table' not in wb.sheetnames:
        raise ValueError("Resource_Table sheet not found in the workbook!")
        
    sheet = wb['Resource_Table']
    # Iterate through rows starting from row 2 (skipping header)
    for row in range(2, sheet.max_row + 1):
        cell_id = sheet.cell(row=row, column=1).value
        if cell_id is not None:
            try:
                rid = int(cell_id)
                if rid in NEW_WAGES:
                    new_rate = f"${NEW_WAGES[rid]:.2f}/h"
                    # Update column J (10th column)
                    sheet.cell(row=row, column=10).value = new_rate
                    print(f"  Excel Row {row}: ID {rid} ({sheet.cell(row=row, column=2).value}) -> {new_rate}")
            except ValueError:
                pass
                
    wb.save(filepath)
    print("Excel file updated and saved successfully.")

def update_csv(filepath):
    print(f"Updating CSV file: {filepath}")
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"CSV file not found: {filepath}")
        
    rows = []
    header = None
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader)
        for row in reader:
            if not row:
                continue
            try:
                rid = int(row[0])
                if rid in NEW_WAGES:
                    new_rate = f"${NEW_WAGES[rid]:.2f}/h"
                    row[9] = new_rate
                    print(f"  CSV: ID {rid} ({row[1]}) -> {new_rate}")
            except ValueError:
                pass
            rows.append(row)
            
    with open(filepath, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)
    print("CSV file updated and saved successfully.")

if __name__ == "__main__":
    xlsx_path = "Sample Commercial Construction Schedule.xlsx"
    csv_path = os.path.join("Schedules_CSV", "Resource_Table.csv")
    
    update_xlsx(xlsx_path)
    print("-" * 40)
    update_csv(csv_path)
